#import pandas as pd
import xml.etree.ElementTree as ET
import os
import re
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from datetime import datetime
from openpyxl.worksheet.pagebreak import Break # Import necessário para quebras de página
from typing import Dict, Any, List
import sys

#definições ------------------------------------------
ARQUIVO_UNITPRO = 'unitpro.xef'
TIPOS_PERMITIDOS = ['WORD', 'BOOL', 'EBOOL', 'INT','UINT']


# Mapeamento completo dos cartões
MODELOS_INFO = {
    "140ACI03000": {"canais": 8,  "prefixo": "%IW"},
    "140AVI03000": {"canais": 8,  "prefixo": "%IW"},
    "140ACO02000": {"canais": 4,  "prefixo": "%MW"},
    "140ACO13000": {"canais": 8,  "prefixo": "%MW"},
    "140ARI03010": {"canais": 8,  "prefixo": "%IW"},
    "140DDI84100": {"canais": 32, "prefixo": "%I"},
    "140DAI54000": {"canais": 16, "prefixo": "%I"},
    "140DAI55300": {"canais": 32, "prefixo": "%I"},
    "140DAO84210": {"canais": 16, "prefixo": "%M"},
    "140DAI74000": {"canais": 16, "prefixo": "%M"},
    "140DDI35300": {"canais": 32, "prefixo": "%I"},
    "140DDO35300": {"canais": 32, "prefixo": "%M"},
    "BMXDDI3202K": {"canais": 32, "prefixo": "%I"},
    "BMXDDO3202K": {"canais": 32, "prefixo": "%M"}
}
MODELOS_EXCECAO = ["140CPS", "140CRA", "140NRP","140NOE"]

class Canal:
    def __init__(self, numero):
        self.numero = numero
        self.nome = "-"
        self.comentario = "-"
        self.endereco = ""  
class Slot:
    def __init__(self, numero, modelo, endereco_base):
        self.numero = numero
        self.modelo = modelo
        self.endereco_base = endereco_base  # Ex: "33" ou "ED_DROP..."
        self.qtd_canais = 0
        self.canais = []

        if not any(self.modelo.startswith(p) for p in MODELOS_EXCECAO):
            info = MODELOS_INFO.get(modelo)
            if info:
                self.qtd_canais = info['canais']
                prefixo = info['prefixo']
                
                for i in range(self.qtd_canais):
                    novo_canal = Canal(i + 1)
                    # Se for numérico, gera o endereço sequencial
                    if str(self.endereco_base).isdigit():
                        num_final = int(self.endereco_base) + i
                        novo_canal.endereco = f"{prefixo}{num_final}"
                    self.canais.append(novo_canal)        
class Drop:
    def __init__(self, numero):
        self.numero = numero
        self.slots = {} # Dicionário {numero_slot: Objeto Slot}

def gerar_matriz_plc(caminho):
    tree = ET.parse(caminho)
    root = tree.getroot()
    
    drops = {} # Dicionário {numero_drop: Objeto Drop}

    for module in root.findall(".//moduleQuantum"):
        try:
            # 1. Pegar o Part Number (Modelo)
            part_item = module.find("partItem")
            modelo = part_item.get("partNumber")

            # 2. Lógica para capturar o ENDEREÇO baseado no IOVision
            vision_type = module.get("IOVision")
            endereco = "Desconhecido"

            if vision_type == "device DDT":
                # Busca o atributo implInstName dentro da tag deviceDDT
                device_ddt = module.find("deviceDDT")
                if device_ddt is not None:
                    endereco = device_ddt.get("implInstName")
            
            else:
            #elif vision_type == "state ram full":
                # Busca o atributo inputRefOffset dentro da tag moduleInfo
                module_info = module.find("moduleInfo")
                if module_info is not None:
                    endereco_input = module_info.get("inputRefOffset")
                    endereco_output = module_info.get("outputRefOffset")
                    endereco = max(endereco_input, endereco_output)
                    
            # 3. Pegar o TopoAddress para localizar Drop/Slot
            equip_info = module.find("equipInfo")
            topo_address = equip_info.get("topoAddress")

            # Regex para extrair DROP e SLOT do endereço \2.X\1.Y
            #match = re.search(r'\\2\.(\d+)\\1\.(\d+)', topo_address)
            match = re.search(r'\\\d+\.(\d+)\\1\.(\d+)', topo_address)
            if match:
                num_drop = int(match.group(1))
                num_slot = int(match.group(2))

                # Adiciona Drop se não existir
                if num_drop not in drops:
                    drops[num_drop] = Drop(num_drop)
                
                # Adiciona Slot ao Drop com o novo parâmetro 'endereco'
                drops[num_drop].slots[num_slot] = Slot(num_slot, modelo, endereco)
                
                print(f"Mapeado: Drop {num_drop}, Slot {num_slot} -> Modelo {modelo} | Endereço: {endereco}")

        except Exception as e:
            print(f"Erro ao processar módulo: {e}")

    return drops

def ler_variaveis_unitpro(caminho_arquivo: str) -> Dict[str, Dict[str, str]]:
    """Retorna um dicionário onde a chave é o NOME da variável."""
    mapa_por_nome = {}
    try:
        tree = ET.parse(caminho_arquivo)
        root = tree.getroot()
    except Exception as e:
        print(f"ERRO: {e}")
        return mapa_por_nome

    for var_element in root.findall('.//variables'):
        nome = var_element.get('name')
        if not nome: continue
        
        tipo = var_element.get('typeName')
        endereco_raw = var_element.get('topologicalAddress')
        
        comentario_elem = var_element.find('comment')
        comentario = comentario_elem.text.strip() if comentario_elem is not None and comentario_elem.text else ""

        if tipo in TIPOS_PERMITIDOS:
            mapa_por_nome[nome] = {
                'nome': nome,
                'comentario': comentario,
                'endereco': endereco_raw, # Mantemos o original aqui (%I00033)
                'tipo': tipo
            }
    return mapa_por_nome


def normalizar_endereco(endereco_str):
    """Converte '%I00033' ou '%I33' em ('%I', 33) para comparação justa."""
    if not endereco_str or not endereco_str.startswith("%"):
        return None
    match = re.match(r'(%[a-zA-Z]+)(\d+)', endereco_str)
    if match:
        prefixo = match.group(1).upper()
        numero = int(match.group(2))
        return (prefixo, numero)
    return None

def preencher_canais_da_matriz(caminho_arquivo, matriz_hardware, mapa_por_nome):
    # Índice auxiliar para State RAM: (Prefixo, Numero) -> Nome
    # Isso resolve o problema de busca rápida por endereço
    indice_endereco = {}
    for nome, dados in mapa_por_nome.items():
        if dados['endereco']:
            norm = normalizar_endereco(dados['endereco'])
            if norm:
                indice_endereco[norm] = nome

    tree = ET.parse(caminho_arquivo)
    root = tree.getroot()

    for drop in matriz_hardware.values():
        for slot in drop.slots.values():
            if not slot.canais: continue

            # SITUAÇÃO 1: STATE RAM (Endereço base numérico)
            if str(slot.endereco_base).isdigit():
                for canal in slot.canais:
                    chave_canal = normalizar_endereco(canal.endereco)
                    nome_encontrado = indice_endereco.get(chave_canal)
                    
                    if nome_encontrado:
                        dados = mapa_por_nome[nome_encontrado]
                        canal.nome = dados['nome']
                        canal.comentario = dados['comentario']

            # SITUAÇÃO 2: DDT (Busca o Alias no XML)
            else:
                var_node = root.find(f".//variables[@name='{slot.endereco_base}']")
                if var_node is not None:
                    for ch_desc in var_node.findall(".//instanceElementDesc"):
                        ch_name = ch_desc.get("name", "")
                        if ch_name.startswith("[") and ch_name.endswith("]"):
                            idx = int(ch_name.strip("[]"))
                            if idx < len(slot.canais):
                                val_node = ch_desc.find(".//instanceElementDesc[@name='VALUE']")
                                if val_node is not None:
                                    alias = val_node.find("attribute[@name='Alias']")
                                    if alias is not None:
                                        # Aqui está o pulo do gato: o Alias é o NOME
                                        slot.canais[idx].nome = alias.get("value")

 



def preencher_comentarios_na_matriz(matriz_hardware, mapa_por_nome):
    contador = 0
    for drop in matriz_hardware.values():
        for slot in drop.slots.values():
            for canal in slot.canais:
                if canal.nome and canal.nome != "-":
                    # Busca direta por NOME no dicionário
                    dados = mapa_por_nome.get(canal.nome.strip())
                    if dados:
                        canal.comentario = dados['comentario']
                        contador += 1
    print(f"Sucesso: {contador} comentários processados.")


def ler_titulo_modelo(caminho_arquivo_xef: str, lista_variaveis_lidas: List[Dict[str, Any]]) -> str:
 
    
    try:
        tree = ET.parse(caminho_arquivo_xef)
        root = tree.getroot()
        # Busca o partItem que está dentro de PLC
        plc_part = root.find(".//PLC/partItem")
        if plc_part is not None:
            MODELO = plc_part.get("family", "Modelo Desconhecido")
    except Exception as e:
        print(f"Erro ao extrair família do PLC: {e}")
        MODELO = "PLC"

    """
    Lê o atributo 'name' da tag contentHeader no arquivo XEF.
    Se o título for "Project", procura por uma tag terminada em '_DCOM' E do tipo 'WORD'
    na lista de variáveis e a utiliza como título.
    """

    # --- 1. Lógica Original de Leitura do Título no XML ---
    
    titulo_lido = 'Projeto_Invalido'
    
    try:
        tree = ET.parse(caminho_arquivo_xef)
        root = tree.getroot()
        header = root.find('contentHeader')
        
        if header is not None:
            # Pega o nome. Se não houver, usa 'Projeto_Sem_Nome'.
            titulo_lido = header.get('name', 'Projeto_Sem_Nome')
        else:
            titulo_lido = 'Projeto_Sem_Header'
            
    except (FileNotFoundError, ET.ParseError):
        # Mantém 'Projeto_Invalido'
        pass
        
    # --- 2. Lógica de Verificação e Substituição para "_DCOM" e tipo "WORD" ---
    
    if titulo_lido == "Project":
        
        print("\nAlerta: Título original encontrado é 'Project'. Buscando fallback '_DCOM' (Tipo WORD)...")
        
        # Procura a primeira variável que atenda a ambas as condições
        for variavel in lista_variaveis_lidas:
            nome_variavel = lista_variaveis_lidas[variavel]['nome']
            #variavel.get('nome', '')
            tipo_variavel = lista_variaveis_lidas[variavel]['tipo']
            #variavel.get('tipo', '')
            
            # Condição A: A tag deve terminar com "_DCOM"
            condicao_dcom = nome_variavel and nome_variavel.endswith('_DCOM')
            
            # Condição B: O tipo deve ser "WORD"
            condicao_word = tipo_variavel == 'WORD'
            
            # Verifica se AMBAS as condições são atendidas
            if condicao_dcom and condicao_word:
                print(f"Substituindo 'Project' pela tag: {nome_variavel}")
                return nome_variavel.removesuffix('_DCOM'), MODELO # Retorna imediatamente o novo título
                
        # 3. Se o loop terminar sem encontrar a tag "_DCOM" tipo "WORD"
        print("Aviso: Nenhuma tag '_DCOM' do tipo 'WORD' foi localizada na lista de variáveis lidas.")
        return titulo_lido, MODELO
        
    else:
        # Se o título original for válido e não for "Project", retorna o que foi lido
        return titulo_lido, MODELO



#----------------------GERAÇÃO DO ARQUIVO EXCEL----------------------------

def gerar_excel(matriz_hardware, titulo_projeto, modelo_plc):
    data_hoje = datetime.now().strftime("%Y-%m-%d")
    nome_arquivo = f"REMOTE_IO_{titulo_projeto.upper()}_{data_hoje}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista de IO"

    # --- CONFIGURAÇÕES DE IMPRESSÃO ---
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0 
    
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # Estilos
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True, size=10)
    center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_aligned = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)

    linha_atual = 1

    for num_drop in sorted(matriz_hardware.keys()):
        obj_drop = matriz_hardware[num_drop]
        for num_slot in sorted(obj_drop.slots.keys()):
            if num_slot < 3:
                continue
            obj_slot = obj_drop.slots[num_slot]

            # --- CABEÇALHO (Linha 1 do Slot) ---
            # Mescla a Coluna 1 (A) e 2 (B) para o texto "VALE"
            ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=2)
            ws.cell(row=linha_atual, column=1, value="VALE").font = header_font
            ws.cell(row=linha_atual, column=3, value=titulo_projeto).font = header_font
            ws.cell(row=linha_atual, column=4, value=f"Modelo\n{modelo_plc}").font = header_font
            ws.cell(row=linha_atual, column=5, value=f"Cartão\n{obj_slot.modelo}").font = header_font
            ws.cell(row=linha_atual, column=6, value=f"Drop\n{num_drop:02d}").font = header_font
            ws.cell(row=linha_atual, column=7, value=f"Slot\n{num_slot:02d}").font
            

            for col in range(1, 8):
                cell = ws.cell(row=linha_atual, column=col)
                cell.alignment = center_aligned
                cell.border = thin_border

            # --- LINHA 2 (Subtítulo e Revisão) ---
            ws.merge_cells(start_row=linha_atual+1, start_column=1, end_row=linha_atual+1, end_column=5)
            ws.cell(row=linha_atual+1, column=1, value="Entradas/Saídas Digitais ou Analógicas").alignment = center_aligned
            ws.merge_cells(start_row=linha_atual+1, start_column=6, end_row=linha_atual+1, end_column=7)
            ws.cell(row=linha_atual+1, column=6, value=f"Revisão: {data_hoje}").alignment = center_aligned
            
            for col in range(1, 8):
                ws.cell(row=linha_atual+1, column=col).border = thin_border

            # --- LINHA 3 (Títulos da Tabela) ---
            ws.cell(row=linha_atual+2, column=1, value="BORNE").font = header_font
            ws.cell(row=linha_atual+2, column=2, value="BIT").font = header_font # Nova Coluna
            ws.cell(row=linha_atual+2, column=3, value="TAG Equipamento").font = header_font
            ws.merge_cells(start_row=linha_atual+2, start_column=4, end_row=linha_atual+2, end_column=7)
            ws.cell(row=linha_atual+2, column=4, value="DESCRIÇÃO / COMENTÁRIO").font = header_font
            
            for col in range(1, 8):
                cell = ws.cell(row=linha_atual+2, column=col)
                cell.alignment = center_aligned
                cell.border = thin_border
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

            # --- CANAIS (Preenchimento) ---
            for i in range(32):
                r_idx = linha_atual + 3 + i
                ws.row_dimensions[r_idx].height = 14.5 # Ajuste para caber no A4 Paisagem
                
                # Borne (Coluna 1)
                ws.cell(row=r_idx, column=1, value=i+1).alignment = center_aligned
                
                # Bit (Coluna 2: Borne - 1)
                ws.cell(row=r_idx, column=2, value=i).alignment = center_aligned
                
                tag = "-"
                coment = "-"
                if i < len(obj_slot.canais):
                    tag = obj_slot.canais[i].nome or "-"
                    coment = obj_slot.canais[i].comentario or "-"

                # Tag (Coluna 3)
                ws.cell(row=r_idx, column=3, value=tag).alignment = center_aligned
                
                # Comentário (Colunas 4 a 7 mescladas)
                ws.merge_cells(start_row=r_idx, start_column=4, end_row=r_idx, end_column=7)
                ws.cell(row=r_idx, column=4, value=coment).alignment = left_aligned

                for col in range(1, 8):
                    ws.cell(row=r_idx, column=col).border = thin_border

            # --- FINALIZAÇÃO DO SLOT ---
            linha_atual += 35 
            ws.row_breaks.append(Break(id=linha_atual-1))

    # Ajuste final de colunas (A=Borne, B=Bit, C=Tag, D...G=Comentário)
    larguras = [10, 10, 30, 21, 21, 21, 21] 
    for i, w in enumerate(larguras):
        ws.column_dimensions[chr(65+i)].width = w

    wb.save(nome_arquivo)
    print(f"Arquivo único gerado: {nome_arquivo}")
#----------------------MAIN----------------------------
if __name__ == "__main__":

    # --- DEFINIÇÃO UNIVERSAL DO CAMINHO BASE ---
    # Essa lógica funciona tanto para o script .py quanto para o executável .exe (frozen)
    if getattr(sys, 'frozen', False):
        # Se estiver rodando como executável (PyInstaller), usa o caminho do binário.
        diretorio_script = os.path.dirname(sys.executable)
    else:
        # Se estiver rodando como script Python (.py), usa o caminho do arquivo de script.
        # É fundamental usar o try-except ou um método robusto para evitar erros ao ser chamado de outro diretório.
        try:
            diretorio_script = os.path.dirname(os.path.abspath(__file__))
        except NameError:
            # Fallback caso __file__ não esteja definido (raro, mas seguro)
            diretorio_script = os.path.getcwd() 

    # --- Configuração de Caminhos ---
    caminho_unitpro = os.path.join(diretorio_script, ARQUIVO_UNITPRO)

    # 1. Leitura e Catalogação das Variáveis
    lista_variaveis_lidas = ler_variaveis_unitpro(caminho_unitpro)

    # 2. Gerar a estrutura a partir do hardware do PLC
    matriz_hardware = gerar_matriz_plc(caminho_unitpro)
 
    # 3. Preencher os nomes dos canais com base nas variáveis do arquivo
    preencher_canais_da_matriz(caminho_unitpro, matriz_hardware,lista_variaveis_lidas)

    # 4. Preencher os COMENTÁRIOS nos canais
    # Cruza os dados da matriz com a lista_variaveis_lidas
    preencher_comentarios_na_matriz(matriz_hardware, lista_variaveis_lidas)

    # Supondo que você extraiu essas informações do XML ou entrada do usuário:
    titulo_projeto, modelo_plc = ler_titulo_modelo(caminho_unitpro,lista_variaveis_lidas) 
    

    # 6. Geração do arquivo com o nome dinâmico: REMOTE_IO_[UC1000CC21]_2025-12-31.xlsx
    gerar_excel(matriz_hardware, titulo_projeto, modelo_plc)
    print("Processamento concluído.") 